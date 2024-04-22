import pandas as pd
import openpyxl
from playwright.sync_api import sync_playwright
from faker import Faker

random = Faker()

# Browse Code Starts
playwright = sync_playwright().start()

# Entire Automation workbook accessing
# ------------------------------- Automation Project Workbook Accessing -------------------------------------------

Workbook_path = "C:\\Users\\LNV-53\\Desktop\\Framework\\Input\\bluesky_project.xlsx"
workbook = openpyxl.load_workbook(Workbook_path)
Locator = workbook["Locator"]  # Locator Sheet accessed
configuration = workbook["Configuration"]  # Configuration Sheet accessed

Object_dict = {}  # Create blank dictionary for Locators

for row in Locator.iter_rows(min_row=2, values_only=True):  # Exclude first row & iterate remaining rows
    ui_element = row[0]
    locator = row[1]
    Object_dict[ui_element] = locator  # Store values in dictionary (key = value) format
# print(Object_dict)
# print("\n")
config = {}  # Create blank dictionary for Configurations
for row in configuration.iter_rows(min_row=2, values_only=True):  # Exclude first row & iterate remaining rows
    param = row[0]
    value = row[1]
    config[param] = value  # Store values in dictionary (key = value) format

#print(config)

# ----------------------------------------------------------------------------------------------------------------
TestSuites = pd.read_excel(Workbook_path, "Test Suites")  # Test Suite Sheet is accessing
# print(TestSuites)
# print("\n")

#############################################################################################################
Filtered_data = TestSuites[TestSuites["Run"] == "Yes"]  # Filtering test suites which having flag 'Yes'
# print(Filtered_data)
############################################################################

# Creating list of test suites which we want to run.
############################################################################
Test_suite_name = []
for Suite_name in Filtered_data["Test Suite Name"]:
    Test_suite_name.append(Suite_name)
print("\n")
############################################################################
for sheet_name in Test_suite_name:
    testCaseSheet = pd.read_excel(Workbook_path, sheet_name=sheet_name)
    print(("\n"))
    # print(testCaseSheet)
    ##########   Calculating total number of rows and columns in test case sheet ###################

    testCaseSheetrows = len(testCaseSheet.axes[0])
    testCaseSheetcols = len(testCaseSheet.axes[1])
    # print(testCaseSheetcols)
    # Finding Total test case count and then particular test case start and end position row wise -
    ############################################################################
    no = 1
    totalTestcase = 0
    TestCaseStart_number = 0
    TestCaseEnd_number = 0

    Test_case_range = {}
    for i, row in testCaseSheet.iterrows():

        if row["Test Case Details"] == "Test Case Start":
            TestCaseStart_number = i
            TestCaseNumber = 'TC' + str(no)
            Test_case_range[TestCaseNumber] = {}
            # Adding elements one at a time
            Test_case_range[TestCaseNumber]['Start Point'] = TestCaseStart_number
            runFlag = testCaseSheet.iloc[i-2, testCaseSheet.columns.get_loc('Run')]
            #print(runFlag)
            Test_case_range[TestCaseNumber]['Run'] = runFlag

        if row["Test Case Details"] == "Test Case End":
            TestCaseEnd_number = i
            Test_case_range[TestCaseNumber]['End Point'] = TestCaseEnd_number
            # print(f"Test case started at {TestCaseStart_number} and ends at row number {TestCaseEnd_number}")
            totalTestcase = totalTestcase + 1
            no = no + 1
            TestData1_row_num = TestCaseStart_number - 1
            TestCaseiteration = 0
            for j in range(testCaseSheetcols):
                # print(j)
                # check if a particular cell is null
                col_number = j + 6
                cell_is_null = testCaseSheet.isnull().iloc[TestData1_row_num, int(col_number)]
                if cell_is_null == True:
                    break
                TestCaseiteration = TestCaseiteration + 1
                # print(f"Test data row number {TestData1_row_num} and column number is {col_number} ")
            # print(f"Exited for loop with TestCaseiteration count {TestCaseiteration}")
            Test_case_range[TestCaseNumber]['TestCaseiteration'] = TestCaseiteration


    print(f"Total test cases for {sheet_name} suite is {totalTestcase}")
    print(Test_case_range)

    for key1, value1 in Test_case_range.items():
        if isinstance(value1, dict):
            for key2, value2 in value1.items():
                # print(key2, ":", value2)
                if key2 == "Start Point":
                    startP = value2
                    # print(startP)
                elif key2 == "End Point":
                    endP = value2 + 1
                    # print(endP)
                elif key2 == "TestCaseiteration":
                    iteration_count = value2
                    # print(iteration_count)
                elif key2 == "Run":
                    runFlag = value2

            if runFlag == "Yes":

                # iterate as per test case iteration count
                for i in range(iteration_count):

                    # print(i)
                    # Browse Code Starts
                    # playwright = sync_playwright().start()
                    browser = playwright.chromium.launch(headless=False, slow_mo=10000)
                    context = browser.new_context(record_video_dir="videos/")
                    page = context.new_page()
                    # Browse Code Ends
                    for x in range(startP, endP):
                        # print(testCaseSheet['Action'].loc[testCaseSheet.index[x]])
                        Action = testCaseSheet['Action'].loc[testCaseSheet.index[x]]
                        Element = testCaseSheet['Object'].loc[testCaseSheet.index[x]]
                        Element = Element.replace(" ", "")
                        Element = Object_dict[Element]
                        # TestData = testCaseSheet['TestData1'].loc[testCaseSheet.index[x]]
                        TestData = testCaseSheet['TestData' + str(i + 1)].loc[testCaseSheet.index[x]]
                        match Action:
                            case "AppLaunch":
                                #print(Element + '||' + Action + '||' + str(TestData))
                                page.goto(config['URL'])

                            case "Enter value in Textbox":
                                #print(Element + '||' + Action + '||' + str(TestData))
                                match TestData:
                                    case "random.name":
                                        page.get_by_placeholder(Element).fill(str(random.name()))
                                    case "random.text":
                                        page.get_by_placeholder(Element).fill(str(random.text()))
                                    case "random.email":
                                        page.get_by_placeholder(Element).fill(str(random.email()))
                                    case "random.country":
                                        page.get_by_placeholder(Element).fill(str(random.country()))
                                    case "random.url":
                                        page.get_by_placeholder(Element).fill(str(random.url()))
                                    case "random.latitude":
                                        page.get_by_placeholder(Element).fill(str(random.latitude()))
                                    case _:
                                        page.get_by_placeholder(Element).fill(str(TestData))

                            case "Click Button":
                                #print(Element + '||' + Action + '||' + str(TestData))
                                page.get_by_role("button", name=Element).click()

                            case "Click Tab":
                                #print(Element + '||' + Action + '||' + str(TestData))
                                page.get_by_role("tab", name=Element).click()

                            case "Select Dropdown":
                                #print(Element + '||' + Action + '||' + str(TestData))
                                page.get_by_placeholder(Element).click()
                                page.get_by_role("option", name=TestData).click()

                            case "Select Radiobutton":
                                page.get_by_label(Element).check()

                            case "Select checkbox":
                                page.get_by_label(Element).check()

                            case "Link Click":
                                print(Element + '||' + Action + '||' + str(TestData))
                                # page.get_by_role("link", name="Contact").click()
                                page.get_by_role("link", name=Element).click()

                            case "Verify Text":
                                #print(Element + '||' + Action + '||' + str(TestData))
                                page.get_by_text(TestData)

                            case "Wait":
                                #print(Element + '||' + Action + '||' + str(TestData))
                                page.wait_for_timeout(5000)

                            case _:
                                print("No Case for this Action")

                    # Make sure to close, so that videos are saved.
                    context.close()
                    browser.close()
                    path = page.video.path()
                    #print(path)
                    print("---------------------------------------------------------------------")
        else:
            print(key1, ":", value1)
